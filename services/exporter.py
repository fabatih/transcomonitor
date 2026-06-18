"""
services/exporter.py — Export profiles for transcomonitor.

Exports the mappings in various formats and profiles, intended for downstream
consumers (PMSI groupage, indicateurs, registres, interop sémantique).

Profiles :
  - complete    : XLSX multi-onglets identical to transcodage_pipeline_complete.xlsx
  - pmsi        : minimal CSV for PMSI groupage (code_cim10, code_cim11_final, classant, cma)
  - audit       : audit_events JSON or CSV (admin only)
  - foundation  : JSON-LD profile exposing foundation URIs (interop SNOMED/MeSH)
  - diff        : diff between two frozen versions

Each profile is a pure function (con, **filters) → bytes + content_type + filename.
"""
from __future__ import annotations

import csv
import io
import json
import sqlite3
from datetime import datetime
from typing import Iterable, Optional


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────

def _now_label() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _fetch_mappings_dict(
    con: sqlite3.Connection, direction: str,
    *, version_id: Optional[int] = None, only_status: Optional[Iterable[str]] = None,
) -> list[dict]:
    sql = """
        SELECT m.*, c.libelle_fr AS cim10_libelle,
               c.est_classant, c.est_cma, c.niveau_cma, c.chapitre AS cim10_chapitre,
               l.label_fr AS cim11_libelle,
               l.chapitre AS cim11_chapitre
        FROM mappings m
        LEFT JOIN cim10_codes c ON c.code = m.source_code
        LEFT JOIN cim11_linearizations l ON l.code = m.source_code
                                          AND l.release = (SELECT version_label FROM nomenclature_versions WHERE id = m.source_version_id)
        WHERE m.direction = ?
    """
    params: list = [direction]
    if version_id is not None:
        sql += " AND m.current_version_id = ?"; params.append(version_id)
    if only_status:
        ph = ",".join("?" for _ in only_status)
        sql += f" AND m.status IN ({ph})"; params.extend(only_status)
    sql += " ORDER BY m.source_code"
    return [dict(r) for r in con.execute(sql, params).fetchall()]


# ─────────────────────────────────────────────────────────────────────────
# Profile: complete (XLSX, 2 sheets : forward + reverse)
# ─────────────────────────────────────────────────────────────────────────

def export_complete_xlsx(
    con: sqlite3.Connection,
    *, version_id: Optional[int] = None,
) -> tuple[bytes, str, str]:
    """Return XLSX bytes with forward + reverse sheets (and metadata)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Forward sheet
    ws_fwd = wb.create_sheet("CIM-10 vers CIM-11")
    forward_cols = [
        "id", "source_code", "cim10_libelle", "target_kind", "target_mms_code",
        "target_foundation_uris", "relation_type", "fiabilite", "source_decision",
        "status", "est_classant", "est_cma", "niveau_cma", "cim10_chapitre",
        "last_validated_at", "updated_at",
    ]
    ws_fwd.append(forward_cols)
    for cell in ws_fwd[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    for m in _fetch_mappings_dict(con, "forward", version_id=version_id):
        ws_fwd.append([m.get(c) for c in forward_cols])

    # Reverse sheet
    ws_rev = wb.create_sheet("CIM-11 vers CIM-10")
    reverse_cols = [
        "id", "source_code", "cim11_libelle", "target_kind", "target_cim10_code",
        "target_foundation_uris", "relation_type", "fiabilite", "source_decision",
        "status", "cim11_chapitre", "last_validated_at", "updated_at",
    ]
    ws_rev.append(reverse_cols)
    for cell in ws_rev[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    for m in _fetch_mappings_dict(con, "reverse", version_id=version_id):
        ws_rev.append([m.get(c) for c in reverse_cols])

    # Metadata sheet
    ws_meta = wb.create_sheet("Métadonnées")
    ws_meta.append(["Clé", "Valeur"])
    ws_meta.append(["Généré le", datetime.now().isoformat(timespec="seconds")])
    if version_id is not None:
        v_row = con.execute(
            "SELECT label, description, frozen_at FROM frozen_versions WHERE id=?",
            (version_id,),
        ).fetchone()
        if v_row:
            ws_meta.append(["Version", v_row[0]])
            ws_meta.append(["Description", v_row[1]])
            ws_meta.append(["Gelée le", v_row[2]])
    ws_meta.append(["Forward count", ws_fwd.max_row - 1])
    ws_meta.append(["Reverse count", ws_rev.max_row - 1])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"transcomonitor_complete_{_now_label()}.xlsx"
    return (buf.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename)


# ─────────────────────────────────────────────────────────────────────────
# Profile: PMSI groupage (minimal CSV)
# ─────────────────────────────────────────────────────────────────────────

def export_pmsi_csv(
    con: sqlite3.Connection,
    *, version_id: Optional[int] = None,
    only_valide: bool = True,
) -> tuple[bytes, str, str]:
    """CSV minimal pour PMSI groupage : code_cim10, code_cim11, classant, cma."""
    statuses = ("valide", "gele") if only_valide else None
    rows = _fetch_mappings_dict(con, "forward", version_id=version_id,
                                 only_status=statuses)

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["code_cim10", "libelle_cim10", "code_cim11_final",
                "fiabilite", "est_classant", "est_cma", "niveau_cma",
                "relation_type", "status"])
    for m in rows:
        w.writerow([
            m["source_code"], m.get("cim10_libelle") or "",
            m.get("target_mms_code") or "",
            m.get("fiabilite") or "", int(m.get("est_classant") or 0),
            int(m.get("est_cma") or 0), m.get("niveau_cma") or "",
            m.get("relation_type") or "", m.get("status") or "",
        ])
    filename = f"transcomonitor_pmsi_{_now_label()}.csv"
    return (buf.getvalue().encode("utf-8-sig"),
            "text/csv; charset=utf-8",
            filename)


# ─────────────────────────────────────────────────────────────────────────
# Profile: foundation (JSON-LD + CSV)
# ─────────────────────────────────────────────────────────────────────────

JSONLD_CONTEXT = {
    "@vocab": "http://transcomonitor.atih.fr/vocab#",
    "skos":    "http://www.w3.org/2004/02/skos/core#",
    "icd":     "http://id.who.int/icd/",
    "id":      "@id",
    "type":    "@type",
    "mappings": {"@container": "@list"},
    "sourceCode":   "skos:notation",
    "sourceLabel":  "skos:prefLabel",
    "foundationURIs": {"@id": "skos:related", "@container": "@set"},
    "mmsCodes":       {"@type": "skos:notation", "@container": "@set"},
    "release":        "skos:editorialNote",
}


def export_foundation_jsonld(
    con: sqlite3.Connection,
    *, version_id: Optional[int] = None,
    only_valide: bool = True,
) -> tuple[bytes, str, str]:
    """JSON-LD profile : each forward mapping → a skos:Concept with foundation URIs."""
    statuses = ("valide", "gele") if only_valide else None
    rows = _fetch_mappings_dict(con, "forward", version_id=version_id,
                                 only_status=statuses)

    items = []
    for m in rows:
        try:
            fnd = json.loads(m["target_foundation_uris"]) if m["target_foundation_uris"] else []
        except (json.JSONDecodeError, TypeError):
            fnd = []
        items.append({
            "id": f"cim10:{m['source_code']}",
            "type": "skos:Concept",
            "sourceCode": m["source_code"],
            "sourceLabel": m.get("cim10_libelle") or "",
            "mmsCodes": [m["target_mms_code"]] if m.get("target_mms_code") else [],
            "foundationURIs": fnd,
            "relation": m.get("relation_type"),
            "fiabilite": m.get("fiabilite"),
            "status": m.get("status"),
        })

    doc = {
        "@context": JSONLD_CONTEXT,
        "@id": "http://transcomonitor.atih.fr/exports/foundation",
        "type": "skos:ConceptScheme",
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "count": len(items),
        "mappings": items,
    }

    filename = f"transcomonitor_foundation_{_now_label()}.jsonld"
    return (json.dumps(doc, ensure_ascii=False, indent=2).encode("utf-8"),
            "application/ld+json",
            filename)


def export_foundation_csv(
    con: sqlite3.Connection,
    *, version_id: Optional[int] = None,
    only_valide: bool = True,
) -> tuple[bytes, str, str]:
    """Flat CSV variant : one row per (mapping × foundation_uri)."""
    statuses = ("valide", "gele") if only_valide else None
    rows = _fetch_mappings_dict(con, "forward", version_id=version_id,
                                 only_status=statuses)

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["cim10_code", "cim10_libelle", "mms_code", "foundation_uri",
                "relation_type", "fiabilite", "status"])
    for m in rows:
        try:
            fnd = json.loads(m["target_foundation_uris"]) if m["target_foundation_uris"] else []
        except (json.JSONDecodeError, TypeError):
            fnd = []
        if fnd:
            for uri in fnd:
                w.writerow([m["source_code"], m.get("cim10_libelle") or "",
                            m.get("target_mms_code") or "", uri,
                            m.get("relation_type") or "",
                            m.get("fiabilite") or "", m.get("status") or ""])
        else:
            # Still emit a row even without foundation URI (preserves universe)
            w.writerow([m["source_code"], m.get("cim10_libelle") or "",
                        m.get("target_mms_code") or "", "",
                        m.get("relation_type") or "",
                        m.get("fiabilite") or "", m.get("status") or ""])
    filename = f"transcomonitor_foundation_{_now_label()}.csv"
    return (buf.getvalue().encode("utf-8-sig"),
            "text/csv; charset=utf-8",
            filename)


# ─────────────────────────────────────────────────────────────────────────
# Profile: audit (admin only — checked at UI level)
# ─────────────────────────────────────────────────────────────────────────

def export_audit_csv(
    con: sqlite3.Connection,
    *, since: Optional[str] = None, until: Optional[str] = None,
) -> tuple[bytes, str, str]:
    where = "1=1"
    params: list = []
    if since:
        where += " AND ts >= ?"; params.append(since)
    if until:
        where += " AND ts <= ?"; params.append(until)
    rows = con.execute(
        f"""SELECT ts, actor_user_id, actor_username, action,
                   object_type, object_id, old_value_json, new_value_json,
                   source, request_ip, request_ua, note
            FROM audit_events WHERE {where}
            ORDER BY ts ASC, id ASC""",
        params,
    ).fetchall()

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["ts", "actor_user_id", "actor_username", "action",
                "object_type", "object_id", "old_value_json", "new_value_json",
                "source", "request_ip", "request_ua", "note"])
    for r in rows:
        w.writerow(list(r))

    filename = f"transcomonitor_audit_{_now_label()}.csv"
    return (buf.getvalue().encode("utf-8-sig"),
            "text/csv; charset=utf-8",
            filename)


# ─────────────────────────────────────────────────────────────────────────
# Profile: diff between two frozen versions
# ─────────────────────────────────────────────────────────────────────────

def export_version_diff_csv(
    con: sqlite3.Connection,
    *, from_version_id: int, to_version_id: int,
) -> tuple[bytes, str, str]:
    """CSV listing mappings that differ between two frozen_versions."""
    # Get snapshots (may be partial — only if mod-version-freeze has been used)
    fr = {r["source_code"]: dict(r) for r in con.execute(
        "SELECT * FROM version_mappings_snapshot WHERE version_id=?", (from_version_id,)
    ).fetchall()}
    to = {r["source_code"]: dict(r) for r in con.execute(
        "SELECT * FROM version_mappings_snapshot WHERE version_id=?", (to_version_id,)
    ).fetchall()}

    all_codes = sorted(set(fr) | set(to))

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(["source_code", "change_type",
                "target_mms_code_from", "target_mms_code_to",
                "relation_type_from", "relation_type_to",
                "status_from", "status_to"])
    for code in all_codes:
        a = fr.get(code)
        b = to.get(code)
        if a is None:
            change = "added"
        elif b is None:
            change = "removed"
        elif (a.get("target_mms_code") != b.get("target_mms_code")
              or a.get("relation_type") != b.get("relation_type")
              or a.get("status_at_freeze") != b.get("status_at_freeze")):
            change = "modified"
        else:
            continue  # no change → skip
        w.writerow([code, change,
                    a.get("target_mms_code") if a else "",
                    b.get("target_mms_code") if b else "",
                    a.get("relation_type") if a else "",
                    b.get("relation_type") if b else "",
                    a.get("status_at_freeze") if a else "",
                    b.get("status_at_freeze") if b else ""])

    filename = f"transcomonitor_diff_{from_version_id}_vs_{to_version_id}_{_now_label()}.csv"
    return (buf.getvalue().encode("utf-8-sig"),
            "text/csv; charset=utf-8",
            filename)
