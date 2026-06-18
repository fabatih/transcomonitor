"""tests/test_ingest.py — Tests for services/ingest.py

Uses a small synthetic XLSX with 3 forward + 2 reverse rows to validate
the full ingest flow deterministically (no dependency on the 10MB seed).
"""
from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from services import ingest

SCHEMA = (ROOT / "db" / "schema_sqlite.sql").read_text(encoding="utf-8")


# ─────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────

@pytest.fixture
def con() -> sqlite3.Connection:
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys=ON")
    c.executescript(SCHEMA)
    yield c
    c.close()


@pytest.fixture
def synthetic_xlsx(tmp_path) -> Path:
    """Build a small synthetic XLSX matching the pipeline schema."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Guide sheet (not used by ingest, but mimics structure)
    g = wb.create_sheet("Guide")
    g.append(["Pipeline unifiée"])

    # Forward sheet
    fwd = wb.create_sheet(ingest.FORWARD_SHEET)
    fwd.append(ingest.FORWARD_COLUMNS)
    # Row 1 : simple mapping
    fwd.append([
        "A00", "Choléra", "1A00", "Choléra", "TRES_HAUTE", "OMS+ANS",
        None, None, None, False, False, False, None,
        "INTERNATIONAL_BASE", None, None,
        "1A00", "OMS+ANS", "TRES_HAUTE", "Equivalent",
        None, None, None, None,
        None, None, None,
        None, None, None,
        None, None, None, None, None, None,
    ])
    # Row 2 : cluster mapping with traceability
    fwd.append([
        "A011", "Paratyphoïde A", "1A08&XN1K5", "Paratyphoïde A & Vibrio", "HAUTE", "LLM_POST_COORD",
        "valide", None, None, True, False, False, None,
        "INTERNATIONAL_BASE", None, None,
        "1A08", "OMS+ANS", "HAUTE", "Subclass",
        "POST_COORD", "1A08&XN1K5", 0.95, "Post-coordination via specifier",
        None, None, None,
        None, None, None,
        None, None, None, None, "1A08&XN1K5", "A",
    ])
    # Row 3 : non-mappable
    fwd.append([
        "Z999", "Code FR sans équivalent", None, None, "NON_RESOLU", "AUCUNE",
        None, None, None, False, False, False, None,
        "FR_ONLY", None, None,
        None, "AUCUNE", "NON_RESOLU", None,
        None, None, None, None,
        None, None, None,
        None, None, None,
        None, None, None, None, None, None,
    ])

    # Reverse sheet
    rev = wb.create_sheet(ingest.REVERSE_SHEET)
    rev.append(ingest.REVERSE_COLUMNS)
    # Row 1 : reverse mapping with full trace
    rev.append([
        "1A00", "Choléra", "01", "A00", "Choléra", "TRES_HAUTE", "LLM_REVERSE_AMELIORE",
        True, True, False, 2.0,
        "A009", "oms_11to10", "MOYENNE", "A00", "DISCORDANT", 50,
        "AMELIORE", "A00", 0.98, "Le code CIM-11 1A00 inclut...",
        "DEPLACEMENT", "Code OMS A009 (classant) → LLM propose A00",
        "P1",
        None, None, None, None,
        "KEEP_CURRENT", None, 1.0, "A00 déjà classant",
    ])
    # Row 2 : reverse non-mappable
    rev.append([
        "8E66", "Concept CIM-11 sans CIM-10", "08", None, None, "NON_RESOLU", "LLM_REVERSE_SANS_EQUIV",
        False, False, False, None,
        None, None, None, None, None, None,
        "SANS_EQUIVALENT", None, 0.85, "Pas d'équivalent CIM-10",
        None, None, "P3",
        None, None, None, None,
        None, None, None, None,
    ])

    path = tmp_path / "synthetic.xlsx"
    wb.save(str(path))
    return path


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────

class TestHelpers:
    def test_to_bool(self):
        assert ingest._to_bool(True) == 1
        assert ingest._to_bool(False) == 0
        assert ingest._to_bool(None) == 0
        assert ingest._to_bool("") == 0
        assert ingest._to_bool("TRUE") == 1
        assert ingest._to_bool("false") == 0
        assert ingest._to_bool(1) == 1

    def test_to_int(self):
        assert ingest._to_int(2) == 2
        assert ingest._to_int(2.0) == 2
        assert ingest._to_int(None) is None
        assert ingest._to_int("") is None
        assert ingest._to_int("abc") is None
        assert ingest._to_int(True) == 1

    def test_to_str_or_none(self):
        assert ingest._to_str_or_none("hello") == "hello"
        assert ingest._to_str_or_none("  hello  ") == "hello"
        assert ingest._to_str_or_none("") is None
        assert ingest._to_str_or_none(None) is None
        assert ingest._to_str_or_none(42) == "42"

    def test_to_float_or_none(self):
        assert ingest._to_float_or_none(0.95) == 0.95
        assert ingest._to_float_or_none("0.5") == 0.5
        assert ingest._to_float_or_none("") is None
        assert ingest._to_float_or_none(None) is None

    def test_detect_target_kind_forward(self):
        assert ingest._detect_target_kind_forward(None) == "non_mappable"
        assert ingest._detect_target_kind_forward("") == "non_mappable"
        assert ingest._detect_target_kind_forward("BA00") == "mms_simple"
        assert ingest._detect_target_kind_forward("BA00&XN8P1") == "mms_cluster"
        assert ingest._detect_target_kind_forward("1G40/1B5Z") == "mms_cluster"

    def test_detect_relation_type(self):
        assert ingest._detect_relation_type("non_mappable", None, None) == "non_mappable"
        assert ingest._detect_relation_type("mms_cluster", "HAUTE", "OMS+ANS") == "composite"
        assert ingest._detect_relation_type("mms_simple", "HAUTE", "OMS+ANS") == "equivalent"
        assert ingest._detect_relation_type("mms_simple", "HERITAGE", "HERITAGE") == "plus_large"
        assert ingest._detect_relation_type("mms_simple", "CONTESTEE", "BIDIR_CONTESTE") == "ambigu"
        assert ingest._detect_relation_type("mms_simple", "HAUTE", "LLM_POST_COORD") == "necessite_postcoord"

    def test_safe_fiabilite(self):
        assert ingest._safe_fiabilite("HAUTE") == "HAUTE"
        assert ingest._safe_fiabilite("haute") == "HAUTE"
        assert ingest._safe_fiabilite(None) is None
        assert ingest._safe_fiabilite("") is None
        assert ingest._safe_fiabilite("INVALID") is None


# ─────────────────────────────────────────────────────────────────────────
# Versions
# ─────────────────────────────────────────────────────────────────────────

class TestVersions:
    def test_ensure_version_creates(self, con):
        vid = ingest._ensure_version(con, "cim10_fr", "2026")
        assert vid > 0
        row = con.execute("SELECT * FROM nomenclature_versions WHERE id=?", (vid,)).fetchone()
        assert row["nomenclature"] == "cim10_fr"
        assert row["version_label"] == "2026"

    def test_ensure_version_idempotent(self, con):
        vid1 = ingest._ensure_version(con, "cim10_fr", "2026")
        vid2 = ingest._ensure_version(con, "cim10_fr", "2026")
        assert vid1 == vid2


# ─────────────────────────────────────────────────────────────────────────
# Ingest cim10_codes
# ─────────────────────────────────────────────────────────────────────────

class TestIngestCim10:
    def test_ingest_inserts_rows(self, con, synthetic_xlsx):
        n = ingest.ingest_cim10_codes(con, synthetic_xlsx)
        assert n == 3
        rows = list(con.execute("SELECT code, libelle_fr, est_classant, type_code FROM cim10_codes ORDER BY code"))
        codes = [r["code"] for r in rows]
        assert codes == ["A00", "A011", "Z999"]
        # est_classant respected
        a011 = next(r for r in rows if r["code"] == "A011")
        assert a011["est_classant"] == 1
        assert a011["type_code"] == "INTERNATIONAL_BASE"
        z999 = next(r for r in rows if r["code"] == "Z999")
        assert z999["type_code"] == "FR_ONLY"

    def test_ingest_chapitre_proper_cim10(self, con, synthetic_xlsx):
        """Chapter should be derived per CIM-10 international structure (01..22)."""
        ingest.ingest_cim10_codes(con, synthetic_xlsx)
        # A00 → A00-B99 → chapter 01
        row = con.execute("SELECT chapitre FROM cim10_codes WHERE code='A00'").fetchone()
        assert row["chapitre"] == "01"
        # A011 → same chapter
        row = con.execute("SELECT chapitre FROM cim10_codes WHERE code='A011'").fetchone()
        assert row["chapitre"] == "01"
        # Z999 → Z00-Z99 → chapter 21
        row = con.execute("SELECT chapitre FROM cim10_codes WHERE code='Z999'").fetchone()
        assert row["chapitre"] == "21"


# ─────────────────────────────────────────────────────────────────────────
# Ingest forward mappings
# ─────────────────────────────────────────────────────────────────────────

class TestIngestForward:
    def test_forward_inserts(self, con, synthetic_xlsx):
        cim10_v = ingest._ensure_version(con, "cim10_fr", "2026")
        cim11_v = ingest._ensure_version(con, "cim11_mms", "2024-01")
        n = ingest.ingest_forward_mappings(con, synthetic_xlsx,
                                            cim10_version_id=cim10_v,
                                            cim11_release_id=cim11_v)
        assert n == 3
        rows = list(con.execute(
            "SELECT source_code, target_kind, target_mms_code, relation_type, fiabilite, status "
            "FROM mappings WHERE direction='forward' ORDER BY source_code"
        ))
        # A00 → mms_simple
        assert rows[0]["source_code"] == "A00"
        assert rows[0]["target_kind"] == "mms_simple"
        assert rows[0]["target_mms_code"] == "1A00"
        assert rows[0]["relation_type"] == "equivalent"
        assert rows[0]["fiabilite"] == "TRES_HAUTE"
        assert rows[0]["status"] == "propose"
        # A011 → mms_cluster
        assert rows[1]["target_kind"] == "mms_cluster"
        assert rows[1]["target_mms_code"] == "1A08&XN1K5"
        assert rows[1]["relation_type"] == "composite"
        # Z999 → non_mappable
        assert rows[2]["target_kind"] == "non_mappable"
        assert rows[2]["target_mms_code"] is None
        assert rows[2]["relation_type"] == "non_mappable"

    def test_pipeline_trace_preserved(self, con, synthetic_xlsx):
        cim10_v = ingest._ensure_version(con, "cim10_fr", "2026")
        cim11_v = ingest._ensure_version(con, "cim11_mms", "2024-01")
        ingest.ingest_forward_mappings(con, synthetic_xlsx,
                                        cim10_version_id=cim10_v,
                                        cim11_release_id=cim11_v)
        row = con.execute(
            "SELECT pipeline_traceability FROM mappings WHERE source_code='A011' AND direction='forward'"
        ).fetchone()
        trace = json.loads(row["pipeline_traceability"])
        assert trace["etape1_code_cim11"] == "1A08"
        assert trace["etape2_verdict_llm"] == "POST_COORD"
        assert trace["etape2_confiance_llm"] == 0.95
        assert "Post-coordination" in trace["etape2_justification_llm"]

    def test_impacts_aval_populated_from_pmsi_flags(self, con, synthetic_xlsx):
        cim10_v = ingest._ensure_version(con, "cim10_fr", "2026")
        cim11_v = ingest._ensure_version(con, "cim11_mms", "2024-01")
        ingest.ingest_forward_mappings(con, synthetic_xlsx,
                                        cim10_version_id=cim10_v,
                                        cim11_release_id=cim11_v)
        # A011 has est_classant=True
        row = con.execute(
            "SELECT impacts_aval FROM mappings WHERE source_code='A011' AND direction='forward'"
        ).fetchone()
        impacts = json.loads(row["impacts_aval"])
        assert impacts["pmsi_classant"] is True
        # A00 has no PMSI flags → impacts is NULL
        row = con.execute(
            "SELECT impacts_aval FROM mappings WHERE source_code='A00' AND direction='forward'"
        ).fetchone()
        assert row["impacts_aval"] is None


# ─────────────────────────────────────────────────────────────────────────
# Ingest reverse mappings
# ─────────────────────────────────────────────────────────────────────────

class TestIngestReverse:
    def test_reverse_inserts(self, con, synthetic_xlsx):
        cim10_v = ingest._ensure_version(con, "cim10_fr", "2026")
        cim11_v = ingest._ensure_version(con, "cim11_mms", "2024-01")
        n = ingest.ingest_reverse_mappings(con, synthetic_xlsx,
                                            cim10_version_id=cim10_v,
                                            cim11_release_id=cim11_v)
        assert n == 2
        rows = list(con.execute(
            "SELECT source_code, source_kind, target_kind, target_cim10_code, target_mms_code "
            "FROM mappings WHERE direction='reverse' ORDER BY source_code"
        ))
        # 1A00 → A00 (cim10_code)
        assert rows[0]["source_code"] == "1A00"
        assert rows[0]["source_kind"] == "mms_code"
        assert rows[0]["target_kind"] == "cim10_code"
        assert rows[0]["target_cim10_code"] == "A00"
        assert rows[0]["target_mms_code"] is None
        # 8E66 → non_mappable
        assert rows[1]["target_kind"] == "non_mappable"
        assert rows[1]["target_cim10_code"] is None


# ─────────────────────────────────────────────────────────────────────────
# Top-level ingest_seed
# ─────────────────────────────────────────────────────────────────────────

class TestIngestSeed:
    def test_full_ingest_creates_frozen_version(self, con, synthetic_xlsx):
        stats = ingest.ingest_seed(con, synthetic_xlsx, progress=False)
        assert stats["skipped"] is False
        assert stats["cim10_codes"] == 3
        assert stats["forward_mappings"] == 3
        assert stats["reverse_mappings"] == 2
        # frozen_version created
        row = con.execute(
            f"SELECT label, is_initial_seed FROM frozen_versions WHERE label='{ingest.SEED_VERSION_LABEL}'"
        ).fetchone()
        assert row is not None
        assert row["is_initial_seed"] == 1

    def test_idempotent(self, con, synthetic_xlsx):
        ingest.ingest_seed(con, synthetic_xlsx, progress=False)
        # Second run should skip
        stats2 = ingest.ingest_seed(con, synthetic_xlsx, progress=False)
        assert stats2["skipped"] is True
        # Still only 5 mappings (3 forward + 2 reverse)
        n = con.execute("SELECT COUNT(*) FROM mappings").fetchone()[0]
        assert n == 5

    def test_force_reset_wipes_and_reingests(self, con, synthetic_xlsx):
        ingest.ingest_seed(con, synthetic_xlsx, progress=False)
        stats = ingest.ingest_seed(con, synthetic_xlsx, force_reset=True, progress=False)
        assert stats["skipped"] is False
        # Should have the SAME counts (wiped then re-ingested)
        n = con.execute("SELECT COUNT(*) FROM mappings").fetchone()[0]
        assert n == 5
        # New frozen_version created (different id)
        n_v = con.execute("SELECT COUNT(*) FROM frozen_versions").fetchone()[0]
        assert n_v == 1  # only one v_pipeline_initial at a time (force_reset deletes it)

    def test_mappings_tagged_with_version(self, con, synthetic_xlsx):
        stats = ingest.ingest_seed(con, synthetic_xlsx, progress=False)
        seed_v_id = stats["seed_version_id"]
        n_untagged = con.execute(
            "SELECT COUNT(*) FROM mappings WHERE current_version_id IS NULL"
        ).fetchone()[0]
        assert n_untagged == 0
        n_tagged = con.execute(
            "SELECT COUNT(*) FROM mappings WHERE current_version_id = ?", (seed_v_id,)
        ).fetchone()[0]
        assert n_tagged == 5

    def test_audit_event_recorded(self, con, synthetic_xlsx):
        ingest.ingest_seed(con, synthetic_xlsx, progress=False)
        rows = list(con.execute(
            "SELECT action, source, object_type, note FROM audit_events"
        ))
        assert len(rows) == 1
        assert rows[0]["action"] == "freeze_version"
        assert rows[0]["source"] == "system"
        assert "Initial seed ingested" in rows[0]["note"]

    def test_missing_file_raises(self, con, tmp_path):
        with pytest.raises(FileNotFoundError):
            ingest.ingest_seed(con, tmp_path / "doesnotexist.xlsx", progress=False)


# ─────────────────────────────────────────────────────────────────────────
# Column validation
# ─────────────────────────────────────────────────────────────────────────

class TestColumnValidation:
    def test_unexpected_columns_raise(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        fwd = wb.create_sheet(ingest.FORWARD_SHEET)
        fwd.append(["wrong_col", "another_wrong"])
        fwd.append(["x", "y"])
        path = tmp_path / "bad.xlsx"
        wb.save(str(path))
        with pytest.raises(ValueError, match="Unexpected columns"):
            list(ingest._iter_sheet(path, ingest.FORWARD_SHEET, ingest.FORWARD_COLUMNS))
