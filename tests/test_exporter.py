"""tests/test_exporter.py — Tests for services/exporter.py"""
from __future__ import annotations

import csv
import io
import json
import sqlite3
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from services import exporter

SCHEMA = (ROOT / "db" / "schema_sqlite.sql").read_text(encoding="utf-8")


@pytest.fixture
def con() -> sqlite3.Connection:
    c = sqlite3.connect(":memory:")
    c.row_factory = sqlite3.Row
    c.execute("PRAGMA foreign_keys=ON")
    c.executescript(SCHEMA)
    # Seed minimal data
    c.execute("INSERT INTO cim10_codes (code, libelle_fr, chapitre, est_classant, est_cma, niveau_cma) VALUES ('A011', 'Typhoid', '01', 1, 1, 2)")
    c.execute("INSERT INTO cim10_codes (code, libelle_fr, chapitre) VALUES ('I10', 'HTA', '09')")
    c.execute("INSERT INTO cim10_codes (code, libelle_fr, chapitre) VALUES ('Z00', 'Examen', '21')")
    # Forward mappings : 2 validés + 1 proposé
    fwd = [
        ("A011", "mms_simple", "1A07.Z", "equivalent", "TRES_HAUTE", "OMS+ANS", "valide",
         json.dumps(["http://id.who.int/icd/entity/1376721186"])),
        ("I10",  "mms_cluster", "BA00&XN8P1", "composite", "HAUTE", "LLM_POST_COORD", "valide",
         json.dumps(["http://id.who.int/icd/entity/588616678",
                     "http://id.who.int/icd/entity/2018344096"])),
        ("Z00",  "non_mappable", None, "non_mappable", "BASSE", "AUCUNE", "propose", None),
    ]
    for code, kind, tgt, rel, fiab, src, status, fnd in fwd:
        c.execute("""INSERT INTO mappings (direction, source_code, source_kind, target_kind,
                                            target_mms_code, target_foundation_uris,
                                            relation_type, fiabilite, source_decision, status)
                     VALUES ('forward', ?, 'cim10_code', ?, ?, ?, ?, ?, ?, ?)""",
                  (code, kind, tgt, fnd, rel, fiab, src, status))
    # Reverse mapping
    c.execute("""INSERT INTO mappings (direction, source_code, source_kind, target_kind,
                                        target_cim10_code, relation_type, fiabilite, source_decision, status)
                 VALUES ('reverse', '1A07.Z', 'mms_code', 'cim10_code', 'A011', 'equivalent',
                         'TRES_HAUTE', 'ALGO_OMS', 'valide')""")
    c.commit()
    yield c
    c.close()


# ─────────────────────────────────────────────────────────────────────────
# Complete XLSX
# ─────────────────────────────────────────────────────────────────────────

class TestCompleteXlsx:
    def test_export_returns_xlsx_bytes(self, con):
        data, ct, fn = exporter.export_complete_xlsx(con)
        assert ct == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert fn.startswith("transcomonitor_complete_") and fn.endswith(".xlsx")
        # Verify XLSX magic bytes (PK\x03\x04)
        assert data[:4] == b"PK\x03\x04"
        assert len(data) > 1000

    def test_xlsx_contains_expected_sheets(self, con):
        import openpyxl
        data, _, _ = exporter.export_complete_xlsx(con)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        names = wb.sheetnames
        assert "CIM-10 vers CIM-11" in names
        assert "CIM-11 vers CIM-10" in names
        assert "Métadonnées" in names

    def test_xlsx_row_counts(self, con):
        import openpyxl
        data, _, _ = exporter.export_complete_xlsx(con)
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
        fwd = wb["CIM-10 vers CIM-11"]
        rev = wb["CIM-11 vers CIM-10"]
        # 3 forward + 1 header
        assert fwd.max_row == 4
        # 1 reverse + 1 header
        assert rev.max_row == 2


# ─────────────────────────────────────────────────────────────────────────
# PMSI CSV
# ─────────────────────────────────────────────────────────────────────────

class TestPmsiCsv:
    def test_default_only_validated(self, con):
        data, ct, fn = exporter.export_pmsi_csv(con)
        assert ct.startswith("text/csv")
        # Decode (BOM-aware via utf-8-sig)
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text), delimiter=";")
        rows = list(reader)
        # Only 2 forward validés (A011, I10) — Z00 propose excluded
        assert len(rows) == 2
        codes = {r["code_cim10"] for r in rows}
        assert codes == {"A011", "I10"}
        a011 = next(r for r in rows if r["code_cim10"] == "A011")
        assert a011["code_cim11_final"] == "1A07.Z"
        assert a011["est_classant"] == "1"
        assert a011["est_cma"] == "1"
        assert a011["niveau_cma"] == "2"

    def test_include_all_with_only_valide_false(self, con):
        data, _, _ = exporter.export_pmsi_csv(con, only_valide=False)
        text = data.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        assert len(rows) == 3  # 2 valide + 1 propose


# ─────────────────────────────────────────────────────────────────────────
# Foundation JSON-LD
# ─────────────────────────────────────────────────────────────────────────

class TestFoundationJsonld:
    def test_returns_valid_jsonld(self, con):
        data, ct, fn = exporter.export_foundation_jsonld(con)
        assert ct == "application/ld+json"
        doc = json.loads(data.decode("utf-8"))
        assert "@context" in doc
        assert doc["type"] == "skos:ConceptScheme"
        assert "mappings" in doc
        assert doc["count"] == 2  # 2 validés
        # Check the structure of one mapping
        m = next(x for x in doc["mappings"] if x["sourceCode"] == "A011")
        assert m["id"] == "cim10:A011"
        assert m["mmsCodes"] == ["1A07.Z"]
        assert m["foundationURIs"] == ["http://id.who.int/icd/entity/1376721186"]

    def test_cluster_has_multiple_foundation_uris(self, con):
        data, _, _ = exporter.export_foundation_jsonld(con)
        doc = json.loads(data.decode("utf-8"))
        m_i10 = next(x for x in doc["mappings"] if x["sourceCode"] == "I10")
        assert len(m_i10["foundationURIs"]) == 2

    def test_context_includes_skos(self, con):
        data, _, _ = exporter.export_foundation_jsonld(con)
        doc = json.loads(data.decode("utf-8"))
        ctx = doc["@context"]
        assert ctx["skos"] == "http://www.w3.org/2004/02/skos/core#"


# ─────────────────────────────────────────────────────────────────────────
# Foundation CSV
# ─────────────────────────────────────────────────────────────────────────

class TestFoundationCsv:
    def test_one_row_per_uri(self, con):
        data, _, _ = exporter.export_foundation_csv(con)
        text = data.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        # A011 has 1 URI → 1 row
        # I10 has 2 URIs → 2 rows
        # Total = 3
        assert len(rows) == 3
        # Verify I10 appears twice with different URIs
        i10_rows = [r for r in rows if r["cim10_code"] == "I10"]
        assert len(i10_rows) == 2
        uris = {r["foundation_uri"] for r in i10_rows}
        assert len(uris) == 2


# ─────────────────────────────────────────────────────────────────────────
# Audit CSV
# ─────────────────────────────────────────────────────────────────────────

class TestAuditCsv:
    def test_export_audit_empty(self, con):
        data, ct, fn = exporter.export_audit_csv(con)
        text = data.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        assert rows == []  # no events yet
        assert ct.startswith("text/csv")

    def test_export_audit_with_events(self, con):
        from services.audit import audit_system_action
        audit_system_action(con, action="cache_refresh", object_type="system", object_id="x")
        audit_system_action(con, action="freeze_version", object_type="system", object_id="y")
        data, _, _ = exporter.export_audit_csv(con)
        text = data.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        assert len(rows) == 2
        actions = [r["action"] for r in rows]
        assert "cache_refresh" in actions
        assert "freeze_version" in actions


# ─────────────────────────────────────────────────────────────────────────
# Version diff
# ─────────────────────────────────────────────────────────────────────────

class TestVersionDiff:
    def test_empty_when_no_snapshots(self, con):
        # Create 2 frozen versions but no snapshots
        con.execute("INSERT INTO frozen_versions (label) VALUES ('v1')")
        con.execute("INSERT INTO frozen_versions (label) VALUES ('v2')")
        con.commit()
        data, _, _ = exporter.export_version_diff_csv(con, from_version_id=1, to_version_id=2)
        text = data.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        assert rows == []

    def test_diff_detects_changes(self, con):
        # Create 2 versions with snapshots
        con.execute("INSERT INTO frozen_versions (label) VALUES ('v1')")
        con.execute("INSERT INTO frozen_versions (label) VALUES ('v2')")
        # v1: A011 → 1A07.Z (mms_simple)
        con.execute("""INSERT INTO version_mappings_snapshot
                       (version_id, mapping_id, direction, source_code, target_kind,
                        target_mms_code, relation_type, status_at_freeze)
                       VALUES (1, 1, 'forward', 'A011', 'mms_simple', '1A07.Z', 'equivalent', 'valide')""")
        # v2: A011 → 1A07.Y (changed)
        con.execute("""INSERT INTO version_mappings_snapshot
                       (version_id, mapping_id, direction, source_code, target_kind,
                        target_mms_code, relation_type, status_at_freeze)
                       VALUES (2, 1, 'forward', 'A011', 'mms_simple', '1A07.Y', 'equivalent', 'valide')""")
        # New code only in v2
        con.execute("""INSERT INTO version_mappings_snapshot
                       (version_id, mapping_id, direction, source_code, target_kind,
                        target_mms_code, relation_type, status_at_freeze)
                       VALUES (2, 2, 'forward', 'I10', 'mms_simple', 'BA00', 'equivalent', 'valide')""")
        # Removed code only in v1
        con.execute("""INSERT INTO version_mappings_snapshot
                       (version_id, mapping_id, direction, source_code, target_kind,
                        target_mms_code, relation_type, status_at_freeze)
                       VALUES (1, 3, 'forward', 'Z00', 'non_mappable', NULL, 'non_mappable', 'rejete')""")
        con.commit()

        data, _, _ = exporter.export_version_diff_csv(con, from_version_id=1, to_version_id=2)
        text = data.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
        assert len(rows) == 3
        changes = {r["source_code"]: r["change_type"] for r in rows}
        assert changes == {"A011": "modified", "I10": "added", "Z00": "removed"}
